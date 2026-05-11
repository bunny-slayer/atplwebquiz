"""Regenerate questions.csv from the local workbook in this folder.

The quiz website loads questions.csv at runtime. If you edit
ATPL_Air_Law_Questions.xlsx, run this script from the atplwebquiz folder:

    python update_questions_from_excel.py

It expects the workbook to contain an "All Questions" sheet with the columns
created by the source parser/exporter.
"""

import csv
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit(
        "Missing dependency: openpyxl\n"
        "Install it with: python -m pip install openpyxl"
    )


ROOT = os.path.dirname(os.path.abspath(__file__))
WORKBOOK_PATH = os.path.join(ROOT, "ATPL_Air_Law_Questions.xlsx")
CSV_PATH = os.path.join(ROOT, "questions.csv")
SHEET_NAME = "All Questions"

CSV_FIELDS = [
    "Section",
    "QuestionID",
    "Question",
    "ChoiceA",
    "ChoiceB",
    "ChoiceC",
    "ChoiceD",
    "CorrectAnswer",
    "Source",
]

HEADER_MAP = {
    "Section": "Section",
    "Question ID": "QuestionID",
    "Question": "Question",
    "Choice A": "ChoiceA",
    "Choice B": "ChoiceB",
    "Choice C": "ChoiceC",
    "Choice D": "ChoiceD",
    "Correct Answer": "CorrectAnswer",
    "Source": "Source",
}


def normalize(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def cell_text(value):
    return "" if value is None else str(value).strip()


def read_rows_from_workbook(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Workbook must contain a "{SHEET_NAME}" sheet.')

    ws = wb[SHEET_NAME]
    header_row = [cell_text(cell.value) for cell in ws[1]]
    col_lookup = {}
    for idx, header in enumerate(header_row):
        if header in HEADER_MAP:
            col_lookup[HEADER_MAP[header]] = idx

    missing = [field for field in CSV_FIELDS if field not in col_lookup]
    if missing:
        raise ValueError(
            "Workbook is missing required columns: " + ", ".join(missing)
        )

    rows = []
    skipped = 0
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {
            field: cell_text(values[col_lookup[field]])
            for field in CSV_FIELDS
        }

        if not row["Question"] or not row["CorrectAnswer"]:
            skipped += 1
            continue

        choices = [row["ChoiceA"], row["ChoiceB"], row["ChoiceC"], row["ChoiceD"]]
        if not any(normalize(choice) == normalize(row["CorrectAnswer"]) for choice in choices):
            row["ChoiceD"] = row["CorrectAnswer"]

        rows.append(row)

    return rows, skipped


def write_csv(rows, path):
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)


def main():
    rows, skipped = read_rows_from_workbook(WORKBOOK_PATH)
    write_csv(rows, CSV_PATH)
    print(f"Wrote {CSV_PATH}")
    print(f"Questions exported: {len(rows)}")
    if skipped:
        print(f"Rows skipped because they were incomplete or invalid: {skipped}")


if __name__ == "__main__":
    main()
